import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_DIR = "outputs/csv"
OUTPUT_XLSX = "outputs/pivot_report.xlsx"

ALL_KD_KAB = [
    9400, 9403, 9408, 9409, 9419, 9420, 9426, 9427, 9428,
    9471,
    9500, 9501, 9502, 9503, 9504,
    9600, 9601, 9602, 9603, 9604, 9605, 9606, 9607, 9608,
    9700, 9701, 9702, 9703, 9704, 9705, 9706, 9707, 9708,
]


def build_assignment_pivot():
    df = pd.read_csv(f"{OUTPUT_DIR}/report_assignment.csv")

    # Only use progress type rows
    df = df[df["type"] == "progress"].copy()

    # Identify SUBMITTED* and COMPLETED*/APPROVED* columns
    status_cols = [c for c in df.columns if c not in (
        "kd_kab", "total", "OPEN", "prov_id", "name", "time_stamp", "type",
        "assigned", "have-not-assigned"
    )]

    submitted_cols = [c for c in status_cols if c.upper().startswith("SUBMITTED")]
    completed_cols = [c for c in status_cols if c.upper().startswith("COMPLETED") or c.upper().startswith("APPROVED") or c.upper().startswith("EDITED")]

    df["Submitted"] = df[submitted_cols].fillna(0).sum(axis=1) if submitted_cols else 0
    df["Completed"] = df[completed_cols].fillna(0).sum(axis=1) if completed_cols else 0
    df["Target"] = df["total"].fillna(0)

    df = df[["name", "kd_kab", "Target", "Submitted", "Completed"]]

    survey_names = df["name"].unique().tolist()
    metrics = ["Target", "Submitted", "Completed"]

    # Build multi-level columns: (survey_name, metric)
    col_tuples = [(s, m) for s in survey_names for m in metrics]
    multi_cols = pd.MultiIndex.from_tuples(col_tuples)

    rows = []
    for kd in ALL_KD_KAB:
        row = {}
        for name in survey_names:
            sub = df[(df["name"] == name) & (df["kd_kab"] == kd)]
            if not sub.empty:
                row[(name, "Target")]    = sub["Target"].iloc[0]
                row[(name, "Submitted")] = sub["Submitted"].iloc[0]
                row[(name, "Completed")] = sub["Completed"].iloc[0]
            else:
                row[(name, "Target")]    = 0
                row[(name, "Submitted")] = 0
                row[(name, "Completed")] = 0
        rows.append(row)

    pivot = pd.DataFrame(rows, index=ALL_KD_KAB, columns=multi_cols)
    pivot.index.name = "Kd Kab"
    return pivot


def build_pemutakhiran_pivot():
    df = pd.read_csv(f"{OUTPUT_DIR}/report_pemutakhiran.csv")

    survey_names = df["name"].unique().tolist()

    # Sub-column order as requested
    sub_cols_ordered = ["total", "BelumSelesaiListing", "SelesaiListing", "SudahTarikSampel"]
    col_map = {c.lower(): c for c in df.columns}
    sub_cols = [col_map.get(c.lower(), c) for c in sub_cols_ordered]
    sub_cols = [c for c in sub_cols if c in df.columns]

    # Build multi-level columns: (survey_name, sub_col)
    col_tuples = [(s, m) for s in survey_names for m in sub_cols]
    multi_cols = pd.MultiIndex.from_tuples(col_tuples)

    rows = []
    for kd in ALL_KD_KAB:
        row = {}
        for name in survey_names:
            sub = df[(df["name"] == name) & (df["kd_kab"] == kd)]
            for m in sub_cols:
                row[(name, m)] = sub[m].iloc[0] if not sub.empty else 0
        rows.append(row)

    pivot = pd.DataFrame(rows, index=ALL_KD_KAB, columns=multi_cols)
    pivot.index.name = "Kd Kab"
    return pivot


def write_pivot_sheet(ws, pivot, sheet_title):
    survey_list = pivot.columns.get_level_values(0).unique().tolist()
    sub_cols    = pivot.columns.get_level_values(1).unique().tolist()
    n_sub       = len(sub_cols)
    n_surveys   = len(survey_list)

    # Styles
    header1_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header2_fill = PatternFill("solid", fgColor="2E75B6")  # mid blue
    index_fill   = PatternFill("solid", fgColor="D6E4F0")  # light blue
    white_font   = Font(bold=True, color="FFFFFF")
    dark_font    = Font(bold=True)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_cols = 1 + n_surveys * n_sub  # kd_kab col + data cols

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=sheet_title)
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1A3657")
    title_cell.alignment = center

    # Row 2: "Kd Kab" header + survey name headers (each merged across n_sub cols)
    kd_cell = ws.cell(row=2, column=1, value="Kd Kab")
    kd_cell.font = white_font
    kd_cell.fill = header1_fill
    kd_cell.alignment = center
    kd_cell.border = border

    for i, survey in enumerate(survey_list):
        start_col = 2 + i * n_sub
        end_col   = start_col + n_sub - 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col, value=survey)
        cell.font = white_font
        cell.fill = header1_fill
        cell.alignment = center
        cell.border = border

    # Row 3: sub-column headers
    ws.cell(row=3, column=1, value="").border = border
    for i in range(n_surveys):
        for j, sub in enumerate(sub_cols):
            col = 2 + i * n_sub + j
            cell = ws.cell(row=3, column=col, value=sub)
            cell.font = white_font
            cell.fill = header2_fill
            cell.alignment = center
            cell.border = border

    # Data rows (one per kd_kab)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for r_idx, (kd, row) in enumerate(pivot.iterrows()):
        excel_row = 4 + r_idx
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        # kd_kab cell
        cell = ws.cell(row=excel_row, column=1, value=kd)
        cell.font = dark_font
        cell.fill = index_fill
        cell.alignment = center
        cell.border = border

        # Data cells
        for i, survey in enumerate(survey_list):
            # Check total/Target column for this survey in this row
            total_col = next((c for c in sub_cols if c in ("total", "Target")), sub_cols[0])
            total_val = row[(survey, total_col)]
            total_numeric = int(total_val) if pd.notna(total_val) else 0

            for j, sub in enumerate(sub_cols):
                col = 2 + i * n_sub + j
                val = row[(survey, sub)]
                numeric = int(val) if pd.notna(val) else 0

                if total_numeric > 0:
                    # Total > 0: show actual value; empty/NaN becomes 0
                    cell_value = numeric
                else:
                    # Total is 0 or empty: leave entire row group empty
                    cell_value = None

                cell = ws.cell(row=excel_row, column=col, value=cell_value)
                cell.fill = row_fill
                cell.alignment = center
                cell.border = border

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 10
    for col in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # Row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 35
    ws.row_dimensions[3].height = 30
    for r in range(4, 4 + len(pivot)):
        ws.row_dimensions[r].height = 18

    # Freeze panes: lock header rows and kd_kab column
    ws.freeze_panes = ws.cell(row=4, column=2)


def main():
    print("Building assignment pivot...")
    assignment_pivot = build_assignment_pivot()

    print("Building pemutakhiran pivot...")
    pemutakhiran_pivot = build_pemutakhiran_pivot()

    print("Writing Excel...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    ws_assign = wb.create_sheet("Assignment")
    write_pivot_sheet(ws_assign, assignment_pivot, "Report Assignment")

    ws_pem = wb.create_sheet("Pemutakhiran")
    write_pivot_sheet(ws_pem, pemutakhiran_pivot, "Report Pemutakhiran")

    wb.save(OUTPUT_XLSX)
    print(f"Saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
